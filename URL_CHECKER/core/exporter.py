"""Export check results to CSV, Excel (openpyxl) or JSON."""

from __future__ import annotations

import csv
import json
import logging
from datetime import datetime
from pathlib import Path

from .models import UrlResult

logger = logging.getLogger(__name__)

COLUMNS = ["url", "status_code", "category", "response_time_ms",
           "final_url", "error", "method", "checked_at"]


def _timestamped_name(fmt: str) -> str:
    ext = {"csv": "csv", "excel": "xlsx", "json": "json"}[fmt]
    return f"url_check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"


def export_results(results: list[UrlResult], fmt: str, out_dir: Path) -> Path:
    """Write results to out_dir in the given format ('csv'|'excel'|'json'). Returns the file path."""
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / _timestamped_name(fmt)
    rows = [r.to_dict() for r in results]

    if fmt == "csv":
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    elif fmt == "json":
        payload = {
            "exported_at": datetime.now().isoformat(timespec="seconds"),
            "total": len(rows),
            "results": rows,
        }
        path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    elif fmt == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        header_fill = PatternFill("solid", fgColor="4F46E5")
        ws.append([c.replace("_", " ").title() for c in COLUMNS])
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        for row in rows:
            ws.append([row.get(c) for c in COLUMNS])
        ws.column_dimensions["A"].width = 70
        ws.freeze_panes = "A2"
        wb.save(path)

    else:
        raise ValueError(f"Unsupported export format: {fmt}")

    logger.info("Exported %d results to %s", len(rows), path)
    return path
